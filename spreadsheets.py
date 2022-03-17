
# You will need to install openpyxl
# pip install openpyxl
# pip3 install openpyxl
# ------------------------
import openpyxl as xl

# Create function to gather which xlsx we are using and which sheets
# ALL ONE FUNCTION
def process_workbook(filename): # ============================================================
    # Create local workbook, which is a copy of the original xlsx file. ======================
    wb = xl.load_workbook(filename) # File in current working directory. spreadsheets.xlsx ===
    sheet = wb["Sheet1"] # Only 1 sheet in this file. This is case sensitive. ================
    #
    # 2 ways to get values of cell ===========================================================
    firstcell = sheet["a1"]
    print(firstcell.value)
    # ----------------------------------------------------------------------------------------
    firstcell = sheet.cell(1, 1)
    print(firstcell.value)
    # -----------------------------------------------------------------------------------------
    #
    #
    print(f'How many active rows in spreadsheet = {sheet.max_row}')
    print(f'How many active columns in spreadsheet = {sheet.max_column}')
    #
    #
    print("\n\nPrint all cells in price column:")
    # range(2   --> start at 2 so so we ignore title cell.====================================
    # range(....... + 1)  --> range does not include last. Need to add 1======================
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)   # Each row other than title row, in column 3==============
        print(cell.value)
        #
        corrected_price = cell.value * 0.9           # Reduce prices by 10% ==================
        corrected_price_cell = sheet.cell(row, 4)    # Stage which cell. Orig-Price = Col-3, corrected price = Col-4
        corrected_price_cell.value = corrected_price # Now actually assign value to cell.=====
    #
    wb.save(filename) # Save document, create new copy doc, don't change original=============
    # ----------------------------------------------------------------------------------------
    # ----------------------------------------------------------------------------------------
    # Now we want to create a chart ==========================================================
    from openpyxl.chart import BarChart, Reference
    #
    # Use new values from column 4 that we just adjusted.=====================================
    values = Reference(sheet,
              min_row=2,                # Avoid row 1, title row.=============================
              max_row=sheet.max_row,    # go to last row holding values=======================
              min_col=4,                # values are only on column 4=========================
              max_col=4                 # values are only on column 4=========================
              )
    #
    chart = BarChart()           # Create BarChart class, and store in our new object.========
    chart.add_data(values)       # Add data to our chart, based on our captured values========
    sheet.add_chart(chart, 'e2') # Which sheet chart will be on, pass our chart, stage chart at cell e2
    #
    wb.save('spreadsheets2.xlsx') # Save document, after chart================================
# -------------------------------------------------
# -------------------------------------------------
# -------------------------------------------------
# We have the single processing xlsx function ALL ABOVE
# Call that fuction with the target file.
# Could easily be automated for many files.
process_workbook("spreadsheets.xlsx")